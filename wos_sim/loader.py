"""Load the data models from "WoS battle simulator.xlsx".

Sources:
  * "Hero Stats" tab   -> HeroRoster (long format: Name, Troop Type, Attribute, Value)
  * "Hero Profile" tab -> HeroProfile per hero, with the avatar image extracted from
    the workbook (pictures are anchored one per hero row in the Photo column)
  * "Troop Stats" tab  -> TroopStats per tier block (pivot blocks labelled e.g.
    "FC9 T11" / "FC10 T11", troop types across columns, attributes down rows)
"""

from __future__ import annotations

import hashlib
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import replace
from pathlib import Path

import openpyxl

from .models import (
    AffectingSide,
    CombatContext,
    DamageCategory,
    EffectReceiver,
    HeroProfile,
    HeroRoster,
    HeroStats,
    PlayerStats,
    SkillAttribute,
    SkillBook,
    SkillCategory,
    SkillEffect,
    SkillMechanic,
    SkillSource,
    StatType,
    TriggerUnit,
    TroopSkillEntry,
    TroopStats,
    TroopType,
)

WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "WoS battle simulator.xlsx"
AVATAR_DIR = Path(__file__).resolve().parent / "data" / "avatars"
HERO_GENERATIONS_PATH = Path(__file__).resolve().parent / "data" / "hero_generations.json"
SKILL_DISPLAY_PATH = Path(__file__).resolve().parent / "data" / "skill_display" / "hero_skills.json"
PROTOTYPE_AVATAR_DIR = Path(__file__).resolve().parent.parent / "prototype" / "avatars"
DEFAULT_TIER = "FC10 T11"
SCRAPED_HERO_STAT = 19.6156

_XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


def load_hero_roster(path: Path = WORKBOOK_PATH) -> HeroRoster:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Hero Stats"]
    grouped: dict[str, dict] = {}
    for name, troop, attr, value in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if name is None:
            continue
        entry = grouped.setdefault(name, {"troop_type": TroopType(troop), "values": {}})
        if entry["troop_type"] != TroopType(troop):
            raise ValueError(f"{name}: inconsistent troop type in Hero Stats tab")
        entry["values"][StatType(attr)] = float(value)
    wb.close()

    roster = HeroRoster()
    for name, entry in grouped.items():
        missing = [s for s in StatType if s not in entry["values"]]
        if missing:
            raise ValueError(f"{name}: missing attributes {missing} in Hero Stats tab")
        roster.add(HeroStats(name=name, troop_type=entry["troop_type"],
                             values=entry["values"]))
    for name, (troop_type, _generation, values) in _SUPPLEMENTAL_HEROES.items():
        if name not in roster:
            roster.add(HeroStats(name=name, troop_type=troop_type,
                                 values=dict(values)))
    for name, meta in _catalog_heroes().items():
        if name not in roster:
            roster.add(HeroStats(
                name=name,
                troop_type=TroopType(meta["troop"]),
                values={stat: SCRAPED_HERO_STAT for stat in StatType},
            ))
    return roster


def load_troop_stats_tiers(path: Path = WORKBOOK_PATH) -> dict[str, TroopStats]:
    """Find every tier block on the "Troop Stats" tab and return them keyed by tier.

    A block starts at a row whose B/C/D cells are the three troop type headers;
    the A cell of that row is the tier label and the next four rows are attributes.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Troop Stats"]
    grid = list(ws.iter_rows(min_row=1, max_row=50, max_col=4, values_only=True))
    wb.close()

    troop_headers = tuple(t.value for t in TroopType)
    tiers: dict[str, TroopStats] = {}
    for i, row in enumerate(grid):
        if tuple(row[1:4]) != troop_headers or row[0] is None:
            continue
        tier = str(row[0])
        table = TroopStats(tier)
        for attr_row in grid[i + 1:i + 1 + len(StatType)]:
            stat = StatType(attr_row[0])
            for col, troop in enumerate(TroopType, start=1):
                table.set(troop, stat, float(attr_row[col]))
        tiers[tier] = table
    return tiers


def load_troop_stats(path: Path = WORKBOOK_PATH, tier: str = DEFAULT_TIER) -> TroopStats:
    tiers = load_troop_stats_tiers(path)
    if tier not in tiers:
        raise KeyError(f"Tier {tier!r} not found; available: {list(tiers)}")
    return tiers[tier]


def load_troop_skill_table(path: Path = WORKBOOK_PATH) -> list[TroopSkillEntry]:
    """Load the skills table on the "Troop Stats" tab (below the tier blocks).

    The table starts at the header row "Troop Type | Skill Name | Attribute |
    Against | Value" and holds Martin's expected-value capture of the innate
    troop skills.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Troop Stats"]
    rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=5, values_only=True))
    wb.close()
    entries: list[TroopSkillEntry] = []
    in_table = False
    for row in rows:
        if row[0] == "Troop Type":
            in_table = True
            continue
        if not in_table:
            continue
        if row[0] is None and row[1] is None:
            if entries:
                break
            continue
        entries.append(TroopSkillEntry(
            troop_type=TroopType(row[0]),
            skill_name=str(row[1]).strip(),
            attribute=None if row[2] is None else str(row[2]).strip(),
            against=None if row[3] is None else str(row[3]).strip(),
            value=float(row[4]),
        ))
    return entries


def _sheet_picture_anchors(z: zipfile.ZipFile, sheet_name: str) -> dict[int, str]:
    """Map 1-based sheet row -> zip path of the picture anchored on that row.

    openpyxl does not expose these pictures, so walk the OOXML relationships:
    workbook -> worksheet -> drawing -> media. Pictures on the "Hero Profile"
    tab are oneCellAnchor'd in the Photo column, one per hero row.
    """
    wbx = ET.fromstring(z.read("xl/workbook.xml"))
    wbrels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid2target = {r.get("Id"): r.get("Target")
                  for r in wbrels.findall("rel:Relationship", _XML_NS)}
    sheet_path = None
    for sh in wbx.findall("main:sheets/main:sheet", _XML_NS):
        if sh.get("name") == sheet_name:
            target = rid2target[sh.get(f"{{{_XML_NS['r']}}}id")].lstrip("/")
            sheet_path = target if target.startswith("xl/") else "xl/" + target
    if sheet_path is None:
        raise KeyError(f"Sheet {sheet_name!r} not found in workbook")

    sheet_rels = ET.fromstring(
        z.read(sheet_path.replace("worksheets/", "worksheets/_rels/") + ".rels"))
    drawing_path = None
    for r in sheet_rels.findall("rel:Relationship", _XML_NS):
        if r.get("Type").endswith("/drawing"):
            target = r.get("Target")
            if target.startswith("/"):
                drawing_path = target.lstrip("/")
            else:
                drawing_path = "xl/" + target.replace("../", "")
            drawing_path = drawing_path.replace("//", "/")
    if drawing_path is None:
        return {}

    drawing_num = re.search(r"drawing(\d+)", drawing_path).group(1)
    drawing_rels = ET.fromstring(z.read(f"xl/drawings/_rels/drawing{drawing_num}.xml.rels"))
    embed2media = {}
    for r in drawing_rels.findall("rel:Relationship", _XML_NS):
        target = r.get("Target")
        if target.startswith("/"):
            media = target.lstrip("/")
        else:
            media = target.replace("../", "xl/")
        embed2media[r.get("Id")] = media.replace("//", "/")

    anchors: dict[int, str] = {}
    drawing = ET.fromstring(z.read(drawing_path))
    for kind in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in drawing.findall(f"xdr:{kind}", _XML_NS):
            blip = anchor.find(".//a:blip", _XML_NS)
            if blip is None:
                continue
            media = embed2media.get(blip.get(f"{{{_XML_NS['r']}}}embed"))
            row0 = int(anchor.find("xdr:from/xdr:row", _XML_NS).text)
            if media:
                anchors[row0 + 1] = media
    return anchors


def _normalize_generation(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _catalog_heroes() -> dict:
    if not HERO_GENERATIONS_PATH.exists():
        return {}
    return json.loads(HERO_GENERATIONS_PATH.read_text(encoding="utf-8"))


def load_hero_profiles(path: Path = WORKBOOK_PATH,
                       avatar_dir: Path = AVATAR_DIR) -> dict[str, HeroProfile]:
    """Load the "Hero Profile" tab and extract each hero's avatar image.

    Avatars are written to avatar_dir as <hero name>.<ext> with a manifest.json
    recording the source media file and content hash per hero.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Hero Profile"]
    rows: dict[int, tuple] = {}
    for idx, (_, name, troop, notes, gen) in enumerate(
            ws.iter_rows(min_row=2, max_col=5, values_only=True), start=2):
        if name is None:
            continue
        rows[idx] = (str(name), TroopType(troop), notes, _normalize_generation(gen))
    wb.close()

    avatar_dir.mkdir(parents=True, exist_ok=True)
    profiles: dict[str, HeroProfile] = {}
    manifest: list[dict] = []
    with zipfile.ZipFile(path) as z:
        anchors = _sheet_picture_anchors(z, "Hero Profile")
        for row, (name, troop, notes, gen) in rows.items():
            media = anchors.get(row)
            avatar_path = None
            if media:
                data = z.read(media)
                avatar_path = avatar_dir / f"{name}{Path(media).suffix.lower()}"
                avatar_path.write_bytes(data)
                manifest.append({
                    "hero": name,
                    "troop_type": str(troop),
                    "generation": gen,
                    "file": avatar_path.name,
                    "source_media": media,
                    "sha1": hashlib.sha1(data).hexdigest(),
                    "bytes": len(data),
                })
            profiles[name] = HeroProfile(name=name, troop_type=troop,
                                         generation=gen, notes=notes,
                                         avatar_path=avatar_path)
    for name, meta in _catalog_heroes().items():
        if name in profiles:
            continue
        avatar_path = None
        if SKILL_DISPLAY_PATH.exists():
            display = json.loads(SKILL_DISPLAY_PATH.read_text(encoding="utf-8"))
            avatar = ((display.get("heroes") or {}).get(name) or {}).get("avatar")
            if avatar:
                candidate = Path(__file__).resolve().parent.parent / "prototype" / avatar
                if candidate.exists():
                    avatar_path = candidate
                    data = candidate.read_bytes()
                    manifest.append({
                        "hero": name,
                        "troop_type": meta["troop"],
                        "generation": str(meta["generation"]),
                        "file": candidate.name,
                        "source_media": str(candidate),
                        "sha1": hashlib.sha1(data).hexdigest(),
                        "bytes": len(data),
                    })
        profiles[name] = HeroProfile(name=name, troop_type=TroopType(meta["troop"]),
                                     generation=str(meta["generation"]),
                                     notes="Scraped from whiteoutsurvival.wiki",
                                     avatar_path=avatar_path)
    for name, (troop, gen, _values) in _SUPPLEMENTAL_HEROES.items():
        profiles.setdefault(
            name,
            HeroProfile(name=name, troop_type=troop, generation=gen,
                        notes="Supplemental Gen-15 profile from official wiki.",
                        avatar_path=None),
        )
    (avatar_dir / "manifest.json").write_text(json.dumps(manifest, indent=2))
    return profiles


# Source-data typo fixes applied on load, keyed by (column header, raw value)
_SKILL_SOURCE_FIXES = {"SKill 3": "Skill 3"}
_HERO_NAME_FIXES = {"Jesse": "Jessie"}  # "Hero Profile" tab spells her Jessie
_SUPPLEMENTAL_HEROES = {
    "Hank": (
        TroopType.INFANTRY,
        "15",
        {StatType.ATTACK: 19.6156, StatType.DEFENSE: 19.6156,
         StatType.LETHALITY: 4.9, StatType.HEALTH: 4.9},
    ),
    "Estrella": (
        TroopType.LANCER,
        "15",
        {StatType.ATTACK: 19.6156, StatType.DEFENSE: 19.6156,
         StatType.LETHALITY: 4.9, StatType.HEALTH: 4.9},
    ),
    "Viveca": (
        TroopType.MARKSMAN,
        "15",
        {StatType.ATTACK: 19.6156, StatType.DEFENSE: 19.6156,
         StatType.LETHALITY: 4.9, StatType.HEALTH: 4.9},
    ),
}
_SUPPLEMENTAL_SKILL_ROWS = [
    *[
        ("Hank", "Skill 1", "All", "Friend", troop, "Lethality", None, 0.25,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Offensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Hank", "Skill 2", "All", "Friend", troop, "Damage Dealt", None, 0.25,
         "Turn-based", 1.0, 0.25, 5.0, "Strikes", None, None, "Both", None,
         "Damage Dealt")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Hank", "Skill 2", "All", "Friend", troop, "Damage Taken", None, -0.25,
         "Turn-based", 1.0, -0.25, 5.0, "Strikes", None, None, "Both", None,
         "Damage Taken")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    ("Hank", "Skill 3", "All", "Foe", "Infantry", "Damage Taken", None, 0.30,
     "Turn-based", 1.0, 0.30, 4.0, "Turns", 2.0, "Turns", "Both", None,
     "Damage Taken"),
    ("Hank", "Skill 3", "All", "Foe", "Marksman", "Damage Dealt", None, -0.30,
     "Turn-based", 1.0, -0.30, 4.0, "Turns", 2.0, "Turns", "Both", None,
     "Damage Dealt"),
    *[
        ("Hank", "Widget", "Garrison", "Friend", troop, "Health", None, 0.15,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Defensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Estrella", "Skill 1", "All", "Foe", troop, "Defense", None, -0.25,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Defensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Estrella", "Skill 2", "All", "Friend", troop, "Attack", None, 0.15,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Offensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Estrella", "Skill 2", "All", "Friend", troop, "Defense", None, 0.10,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Defensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    ("Estrella", "Skill 3", "All", "Friend", "Infantry", "Damage Taken", None, -0.25,
     "Stats-based", None, None, None, None, None, None, "Both", None, "Damage Taken"),
    ("Estrella", "Skill 3", "All", "Friend", "Lancer", "Damage Dealt", None, 0.25,
     "Stats-based", None, None, None, None, None, None, "Both", None, "Damage Dealt"),
    *[
        ("Estrella", "Widget", "Garrison", "Friend", troop, "Attack", None, 0.15,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Offensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    *[
        ("Viveca", "Skill 1", "All", "Friend", troop, "Attack", None, 0.25,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Offensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
    ("Viveca", "Skill 2", "All", "Friend", "Marksman", "Damage Dealt", None, 0.20,
     "Chance-based", 0.20, 1.0, None, None, None, None, "Skills", None, "Damage Dealt"),
    ("Viveca", "Skill 3", "All", "Friend", "Infantry", "Damage Taken", None, -0.10,
     "Stats-based", None, None, None, None, None, None, "Both", None, "Damage Taken"),
    ("Viveca", "Skill 3", "All", "Friend", "Marksman", "Damage Dealt", None, 0.10,
     "Stats-based", None, None, None, None, None, None, "Both", None, "Damage Dealt"),
    *[
        ("Viveca", "Widget", "Rally", "Friend", troop, "Lethality", None, 0.15,
         "Stats-based", None, None, None, None, None, None, "Both", None, "Offensive")
        for troop in ("Infantry", "Lancer", "Marksman")
    ],
]


def _infer_skill_category(attribute: str) -> str:
    if attribute in ("Attack", "Crit Rate", "Lethality"):
        return "Offensive"
    if attribute in ("Defense", "Health"):
        return "Defensive"
    return attribute


def _signed_per_proc(amount, per_proc):
    if per_proc is None:
        return None
    amount = float(amount)
    per_proc = float(per_proc)
    if amount < 0 < per_proc:
        return -per_proc
    if amount > 0 > per_proc:
        return abs(per_proc)
    return per_proc


def _normalize_skill_effect(effect: SkillEffect) -> SkillEffect:
    """Runtime source fixes for workbook rows whose raw sign/side is stale.

    These are intentionally narrow and source-text driven. The workbook remains
    the primary catalog; this layer prevents known stale rows from inverting a
    debuff/buff while the spreadsheet catches up.
    """
    e = replace(effect, amount_per_proc=_signed_per_proc(
        effect.amount, effect.amount_per_proc))
    key = (e.hero, e.source, e.receiver, e.attribute)

    if e.hero == "Cara" and e.source == SkillSource.SKILL_1:
        e = replace(e, amount=-abs(e.amount), amount_per_proc=None)

    if (e.hero == "Vulcanus" and e.source == SkillSource.SKILL_3
            and e.side == AffectingSide.FOE
            and e.attribute == SkillAttribute.DEFENSE):
        e = replace(e, amount=-abs(e.amount),
                    amount_per_proc=-abs(e.amount_per_proc or e.amount))

    if (e.hero == "Flora" and e.source == SkillSource.SKILL_3
            and e.receiver == EffectReceiver.MARKSMAN
            and e.attribute == SkillAttribute.DAMAGE_DEALT):
        e = replace(e, amount=-abs(e.amount),
                    amount_per_proc=-abs(e.amount_per_proc or e.amount))

    if e.hero == "Jeronimo" and e.source == SkillSource.SKILL_3:
        e = replace(e, side=AffectingSide.FRIEND, amount=0.30,
                    amount_per_proc=0.30)

    if e.hero == "Hector" and e.source == SkillSource.SKILL_1:
        e = replace(e, amount=-abs(e.amount),
                    amount_per_proc=-abs(e.amount_per_proc or e.amount))

    if (e.hero == "Nora" and e.source == SkillSource.SKILL_3
            and e.attribute == SkillAttribute.DAMAGE_TAKEN):
        e = replace(e, amount=-abs(e.amount),
                    amount_per_proc=-abs(e.amount_per_proc or e.amount),
                    trigger_unit=TriggerUnit.STRIKES)
    if (e.hero == "Nora" and e.source == SkillSource.SKILL_3
            and e.frequency is not None and e.trigger_unit is None):
        e = replace(e, trigger_unit=TriggerUnit.STRIKES)

    if e.hero == "Philly" and e.source == SkillSource.SKILL_3:
        e = replace(
            e,
            side=AffectingSide.FRIEND,
            attribute=SkillAttribute.DAMAGE_TAKEN,
            amount=-0.20,
            mechanic=SkillMechanic.CHANCE_BASED,
            probability=0.40,
            amount_per_proc=-0.50,
            frequency=None,
            trigger_unit=None,
            duration=None,
            duration_unit=None,
            category=SkillCategory.DAMAGE_TAKEN,
        )

    if e.hero == "Alonso" and e.source == SkillSource.SKILL_1:
        e = replace(
            e,
            side=AffectingSide.FRIEND,
            attribute=SkillAttribute.LETHALITY,
            amount=0.20,
            probability=0.40,
            amount_per_proc=0.50,
            category=SkillCategory.OFFENSIVE,
        )

    if e.hero == "Alonso" and e.source == SkillSource.SKILL_2:
        e = replace(e, amount=-abs(e.amount),
                    amount_per_proc=-abs(e.amount_per_proc or e.amount),
                    duration_unit=TriggerUnit.TURNS,
                    category=SkillCategory.DAMAGE_DEALT)

    if e.hero == "Lynn" and e.source == SkillSource.SKILL_3:
        e = replace(e, trigger_unit=TriggerUnit.STRIKES,
                    duration=None, duration_unit=None)

    if e.hero == "Wayne" and e.source == SkillSource.SKILL_2:
        e = replace(e, trigger_unit=TriggerUnit.STRIKES)

    return e


def _skill_effect_from_values(row) -> SkillEffect:
    (name, source, context, side, receiver, attribute, target, amount,
     mechanic, probability, per_proc, frequency, trigger_unit, duration,
     duration_unit, dmg_category, multiplier, category) = row
    name = _HERO_NAME_FIXES.get(str(name).strip(), str(name).strip())
    source = _SKILL_SOURCE_FIXES.get(str(source).strip(), str(source).strip())
    attribute = str(attribute).strip()
    category = category or _infer_skill_category(attribute)
    return _normalize_skill_effect(SkillEffect(
        hero=name,
        source=SkillSource(source),
        context=CombatContext(context),
        side=AffectingSide(side),
        receiver=EffectReceiver(receiver),
        attribute=SkillAttribute(attribute),
        amount=float(amount),
        mechanic=SkillMechanic(mechanic),
        damage_category=DamageCategory(dmg_category),
        category=SkillCategory(category),
        specific_target=target,
        probability=None if probability is None else float(probability),
        amount_per_proc=None if per_proc is None else float(per_proc),
        frequency=None if frequency is None else float(frequency),
        trigger_unit=None if trigger_unit is None
                     else TriggerUnit(str(trigger_unit).strip()),
        duration=None if duration is None else float(duration),
        duration_unit=None if duration_unit is None
                      else TriggerUnit(str(duration_unit).strip()),
        multiplier=multiplier,
    ))


def _add_display_only_skills(book: SkillBook) -> None:
    if not SKILL_DISPLAY_PATH.exists():
        return
    display = json.loads(SKILL_DISPLAY_PATH.read_text(encoding="utf-8"))
    for hero, record in (display.get("heroes") or {}).items():
        if book.for_hero(hero):
            continue
        troop = TroopType((_catalog_heroes().get(hero) or {}).get("troop", "Infantry"))
        receiver = EffectReceiver(troop.value)
        slots = [SkillSource.SKILL_1, SkillSource.SKILL_2, SkillSource.SKILL_3, SkillSource.WIDGET]
        for source in slots:
            book.add(SkillEffect(
                hero=hero,
                source=source,
                context=CombatContext.ALL,
                side=AffectingSide.FRIEND,
                receiver=receiver,
                attribute=SkillAttribute.ATTACK,
                amount=0.0,
                mechanic=SkillMechanic.STATS_BASED,
                damage_category=DamageCategory.BOTH,
                category=SkillCategory.OFFENSIVE,
                specific_target=None,
            ))


_SKILL_BOOK_CACHE: dict = {}


def _skill_book_rows(path: Path) -> list:
    """Raw "Hero Skills" rows (one atomic effect per row)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return list(wb["Hero Skills"].iter_rows(min_row=2, max_col=18, values_only=True))
    finally:
        wb.close()


def load_skill_book(path: Path | str = WORKBOOK_PATH) -> SkillBook:
    """Load the "Hero Skills" tab: one SkillEffect per row.

    Each row is one atomic effect line; a whole skill (e.g. Dominic's "Mystic
    Mechanism") is the group of rows sharing (hero, skill source).

    The fully-built SkillBook is cached per (path, mtime): the workbook is parsed
    AND the catalog rebuilt once, then the SAME instance is reused across every
    sim in a batch (the turn engine calls this once per run — its hot path).
    Reloaded automatically when the workbook changes on disk, so edits are never
    stale. The SkillBook is a read-only catalog; per-run mutable state (skill
    trigger/kill counters) lives in the SkillDefs built downstream, not here, so
    one shared instance is safe.
    """
    path = Path(path)
    key = (str(path), path.stat().st_mtime_ns)
    cached = _SKILL_BOOK_CACHE.get(key)
    if cached is not None:
        return cached
    book = SkillBook()
    for row in _skill_book_rows(path):
        (name, source, context, side, receiver, attribute, target, amount,
         mechanic, probability, per_proc, frequency, trigger_unit, duration,
         duration_unit, dmg_category, multiplier, category) = row
        if name is None:
            continue
        name = _HERO_NAME_FIXES.get(str(name).strip(), str(name).strip())
        source = _SKILL_SOURCE_FIXES.get(str(source).strip(), str(source).strip())
        category = category or _infer_skill_category(str(attribute).strip())
        book.add(_normalize_skill_effect(SkillEffect(
            hero=name,
            source=SkillSource(source),
            context=CombatContext(context),
            side=AffectingSide(side),
            receiver=EffectReceiver(receiver),
            attribute=SkillAttribute(attribute),
            amount=float(amount),
            mechanic=SkillMechanic(mechanic),
            damage_category=DamageCategory(dmg_category),
            category=SkillCategory(category),
            specific_target=target,
            probability=None if probability is None else float(probability),
            amount_per_proc=None if per_proc is None else float(per_proc),
            frequency=None if frequency is None else float(frequency),
            trigger_unit=None if trigger_unit is None
                         else TriggerUnit(str(trigger_unit).strip()),
            duration=None if duration is None else float(duration),
            duration_unit=None if duration_unit is None
                          else TriggerUnit(str(duration_unit).strip()),
            multiplier=multiplier,
        )))
    loaded = set(book.heroes())
    for row in _SUPPLEMENTAL_SKILL_ROWS:
        if row[0] not in loaded:
            book.add(_skill_effect_from_values(row))
    _add_display_only_skills(book)
    _SKILL_BOOK_CACHE.clear()              # keep only the latest workbook version
    _SKILL_BOOK_CACHE[key] = book
    return book


def new_player_stats(player: str) -> PlayerStats:
    """Empty 12-parameter stat table linked to a player ("Me" / "Enemy")."""
    return PlayerStats(player)


class GameData:
    """Everything the simulator needs, loaded in one call."""

    def __init__(self, path: Path = WORKBOOK_PATH, tier: str = DEFAULT_TIER):
        self.heroes = load_hero_roster(path)
        self.hero_profiles = load_hero_profiles(path)
        self.skills = load_skill_book(path)
        self.troop_skill_table = load_troop_skill_table(path)
        self.troop_stats_tiers = load_troop_stats_tiers(path)
        self.troop_stats = self.troop_stats_tiers[tier]
        self.my_stats = new_player_stats("Me")
        self.enemy_stats = new_player_stats("Enemy")
